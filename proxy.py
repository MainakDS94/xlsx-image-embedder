"""
Local proxy + static server for the Image Embedder HTML app.

Why this exists:
  Browsers block cross-origin image requests from a local HTML file to
  markify.com, even when you're logged in. This tiny server:
    1. Logs into Markify once via Playwright (persistent session).
    2. Serves the HTML app at http://localhost:8765/
    3. Exposes /fetch?url=... which downloads any Markify image using
       your logged-in session and streams the bytes back to the browser,
       with CORS wide open so the HTML can read it.

Setup (one time):
    pip install playwright aiohttp
    playwright install chromium
    python proxy.py --login     # opens browser, log in, press Enter

Run:
    python proxy.py
    # then open http://localhost:8765/ in any browser

Flags:
    --port 8765        change port
    --login            log in and save session, then exit
    --headless         run the background browser headless
"""
import argparse
import asyncio
import base64
import io
import json
import sys
import zipfile
from pathlib import Path

from aiohttp import web
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from playwright.async_api import async_playwright

SESSION_DIR = Path("./markify_session").resolve()
LOGIN_URL = "https://www.markify.com/"
HTML_FILE = Path(__file__).parent / "image_embedder.html"


async def do_login():
    SESSION_DIR.mkdir(exist_ok=True)
    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(
            str(SESSION_DIR), headless=False
        )
        page = await ctx.new_page()
        await page.goto(LOGIN_URL)
        print(f"\nBrowser opened at {LOGIN_URL}")
        print("Log in to Markify, then come back to this terminal.")
        await asyncio.get_event_loop().run_in_executor(
            None, input, "Press Enter once you're logged in to save the session..."
        )
        await ctx.close()
    print(f"Session saved to {SESSION_DIR}")


async def serve(args):
    if not SESSION_DIR.exists():
        sys.exit("No session. Run `python proxy.py --login` first.")
    if not HTML_FILE.exists():
        sys.exit(f"Cannot find {HTML_FILE}. "
                 f"Keep proxy.py and image_embedder.html in the same folder.")

    pw = await async_playwright().start()
    ctx = await pw.chromium.launch_persistent_context(
        str(SESSION_DIR), headless=args.headless
    )
    print(f"Playwright context ready (headless={args.headless}).")

    async def index(request):
        return web.FileResponse(HTML_FILE)

    async def fetch(request):
        url = request.query.get("url")
        if not url:
            return web.Response(status=400, text="Missing ?url=")
        if not url.startswith(("http://", "https://")):
            return web.Response(status=400, text="Bad URL")
        try:
            resp = await ctx.request.get(url, timeout=20000)
            body = await resp.body()
            headers = {
                "Access-Control-Allow-Origin": "*",
                "Content-Type": resp.headers.get(
                    "content-type", "application/octet-stream"
                ),
                "X-Upstream-Status": str(resp.status),
            }
            return web.Response(
                status=resp.status, body=body, headers=headers
            )
        except Exception as e:
            return web.Response(
                status=502,
                text=f"{type(e).__name__}: {e}",
                headers={"Access-Control-Allow-Origin": "*"},
            )

    async def health(request):
        return web.json_response({"ok": True})

    async def build_xlsx(request):
        """
        Build the output xlsx: for each successfully downloaded image, clear
        the URL from its original cell and embed the image there instead.
        Rows where the download failed keep their URL text.

        Expects multipart/form-data with:
          xlsx      (file) - original workbook
          manifest  (text) - JSON {thumb:int, images:[{row:int, col:str, key:str, cellCol:int}]}
                             cellCol is the 1-based column index of the URL cell
          image     (file, repeated) - PNG bytes; part filename = key
        """
        try:
            reader = await request.multipart()
            original_bytes = None
            manifest = None
            images = {}

            while True:
                part = await reader.next()
                if part is None:
                    break
                if part.name == "xlsx":
                    original_bytes = await part.read(decode=False)
                elif part.name == "manifest":
                    manifest = json.loads((await part.read()).decode("utf-8"))
                elif part.name == "image":
                    images[part.filename] = await part.read(decode=False)

            if not original_bytes or manifest is None:
                return web.json_response(
                    {"ok": False, "error": "missing xlsx or manifest"},
                    status=400,
                    headers={"Access-Control-Allow-Origin": "*"},
                )

            thumb = int(manifest.get("thumb", 80))
            placements = manifest.get("images", [])

            wb = load_workbook(io.BytesIO(original_bytes))
            ws = wb.active

            # Widen the columns that will contain images, and set row heights
            # for rows that will get at least one image.
            cols_to_resize = set()
            rows_to_resize = set()
            for p in placements:
                if p["key"] in images:
                    cols_to_resize.add(p["cellCol"])
                    rows_to_resize.add(p["row"])

            for c in cols_to_resize:
                letter = get_column_letter(c)
                current = ws.column_dimensions[letter].width or 8.43
                ws.column_dimensions[letter].width = max(current, thumb / 6)

            for r in rows_to_resize:
                ws.row_dimensions[r].height = thumb * 0.78

            embedded = 0
            skipped = 0
            for p in placements:
                png = images.get(p["key"])
                if not png or len(png) < 8 or png[:8] != b"\x89PNG\r\n\x1a\n":
                    skipped += 1
                    continue
                try:
                    # Clear the URL text so the embedded image is what shows
                    ws.cell(row=p["row"], column=p["cellCol"]).value = None
                    xl_img = XLImage(io.BytesIO(png))
                    xl_img.width = xl_img.height = thumb
                    cell = f"{get_column_letter(p['cellCol'])}{p['row']}"
                    ws.add_image(xl_img, cell)
                    embedded += 1
                except Exception as e:
                    skipped += 1
                    print(f"  skip {p['key']}: {e}")

            out = io.BytesIO()
            wb.save(out)
            body = out.getvalue()
            print(f"build-xlsx: {embedded} embedded, {skipped} skipped, "
                  f"{len(body)/1024/1024:.1f} MB out")
            return web.Response(
                body=body,
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "X-Embedded-Count": str(embedded),
                    "X-Skipped-Count": str(skipped),
                },
            )
        except Exception as e:
            import traceback; traceback.print_exc()
            return web.json_response(
                {"ok": False, "error": f"{type(e).__name__}: {e}"},
                status=500,
                headers={"Access-Control-Allow-Origin": "*"},
            )

    app = web.Application(client_max_size=200 * 1024 * 1024)  # 200 MB for big batches
    app.router.add_get("/", index)
    app.router.add_get("/fetch", fetch)
    app.router.add_get("/health", health)
    app.router.add_post("/build-xlsx", build_xlsx)

    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "127.0.0.1", args.port)
    await site.start()

    print(f"\n  ➜  Open  http://localhost:{args.port}/\n")
    print("Ctrl+C to stop.")
    try:
        await asyncio.Event().wait()  # run forever
    finally:
        await ctx.close()
        await pw.stop()
        await runner.cleanup()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--port", type=int, default=8765)
    ap.add_argument("--login", action="store_true",
                    help="Log in once and save the session, then exit")
    ap.add_argument("--headless", action="store_true",
                    help="Run the background browser headless while serving")
    args = ap.parse_args()

    if args.login:
        asyncio.run(do_login())
        return
    asyncio.run(serve(args))


if __name__ == "__main__":
    main()
